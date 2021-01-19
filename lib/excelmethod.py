from openpyxl import load_workbook

class excelOpern():
    def __init__(self,excelfile_path,sheet_name):
        '''
        获取指定的文件指定sheet的内容,若文件中数据为空则控制台打印错误信息
        :param excelfile_path:
        :param sheet_name:
        '''
        self.wb=load_workbook(excelfile_path).get_sheet_by_name(sheet_name)
        self.sheetname=sheet_name
    def read_excel_rows_dict(self):
        '''读取Excel 文件不存在则返回报错信息，循环读取行信息与第一行组成jsonarray返回，如果单元格没有数据则返回NONE'''
        max_rows=self.wb.max_row
        max_column=self.wb.max_column
        read_list=[]
        title_list=[]
        if max_rows:

            for i in range(0,max_rows):
                rowjson = {}
                for j in range(0,max_column):

                    if i==0:
                        read_list.append(self.wb.cell(row=i+1, column=j+1).value)
                    else:

                        rowjson[read_list[j]]=self.wb.cell(row=i+1, column=j+1).value
                title_list.append(rowjson)
            return title_list[1:]

        else:
            return u"文件里没有数据"